#!/usr/bin/env python3
"""
Complete Flask web app for Competitor Web UI UX Analysis - Step 3
Production Ready for Railway.com
"""

import json
import re
import time
import traceback
import os
from typing import Dict, List, Optional, Generator
from bs4 import BeautifulSoup
from datetime import datetime
import io

from flask import Flask, render_template, request, Response, stream_with_context, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    ElementClickInterceptedException,
    StaleElementReferenceException,
)
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

RATEMYSITE_URL = "https://www.ratemysite.xyz/"
DEFAULT_TIMEOUT = 45

# Global storage for analysis results
analysis_results = {}

def _find_first(driver, xpaths: List[str]) -> Optional[object]:
    for xp in xpaths:
        try:
            el = driver.find_element(By.XPATH, xp)
            if el and el.is_displayed():
                return el
        except (NoSuchElementException, StaleElementReferenceException):
            continue
    return None

def _click_best_button(driver) -> bool:
    xpaths = [
        "//button[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'),'analy')]",
        "//button[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'),'rate')]",
        "//button[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'),'submit')]",
        "//button[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'),'generate')]",
        "//button[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'),'get report')]",
        "//button[@type='submit']",
        "//button",
        "//div[@role='button']",
    ]
    btn = _find_first(driver, xpaths)
    if not btn:
        return False
    try:
        if btn.is_enabled():
            try:
                btn.click()
            except ElementClickInterceptedException:
                driver.execute_script("arguments[0].click();", btn)
            return True
    except Exception:
        pass
    return False

def _maybe_close_cookie_banner(driver):
    candidates = [
        "//button[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'accept')]",
        "//button[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'agree')]",
        "//button[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'allow')]",
        "//button[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'ok')]",
        "//*[contains(@class,'cookie')]//button",
        "//*[@id='onetrust-accept-btn-handler']",
    ]
    try:
        btn = _find_first(driver, candidates)
        if btn:
            try:
                btn.click()
            except ElementClickInterceptedException:
                driver.execute_script("arguments[0].click();", btn)
            time.sleep(0.3)
    except Exception:
        pass

def _collect_result_html(driver) -> str:
    """Get the full HTML of the page for better parsing"""
    try:
        return driver.page_source
    except Exception:
        return ""

def _wait_for_content_growth(driver, wait: WebDriverWait, min_growth: int = 80) -> None:
    try:
        initial_len = len(driver.find_element(By.TAG_NAME, "body").text)
    except Exception:
        initial_len = 0
    try:
        wait.until(lambda d: len(d.find_element(By.TAG_NAME, "body").text) > initial_len + min_growth)
    except TimeoutException:
        pass

def _make_driver(headless: bool = True) -> webdriver.Chrome:
    chrome_opts = Options()
    
    # Production Chrome options for Railway
    chrome_opts.add_argument("--headless=new")
    chrome_opts.add_argument("--no-sandbox")
    chrome_opts.add_argument("--disable-dev-shm-usage")
    chrome_opts.add_argument("--disable-gpu")
    chrome_opts.add_argument("--disable-features=VizDisplayCompositor")
    chrome_opts.add_argument("--window-size=1920,1080")
    chrome_opts.add_argument("--remote-debugging-port=9222")
    chrome_opts.add_argument("--disable-extensions")
    chrome_opts.add_argument("--disable-plugins")
    chrome_opts.add_argument("--disable-images")
    chrome_opts.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
    
    try:
        # Try to use system Chrome first (for Railway)
        service = Service()
        return webdriver.Chrome(service=service, options=chrome_opts)
    except Exception:
        # Fallback to webdriver-manager
        try:
            service = Service(ChromeDriverManager().install())
            return webdriver.Chrome(service=service, options=chrome_opts)
        except Exception as e:
            print(f"Chrome driver setup failed: {e}")
            raise

def _analyze_one_with_debugging(target_url: str, timeout: int = DEFAULT_TIMEOUT) -> tuple[str, List[str]]:
    debug_log = []
    driver = None
    
    try:
        debug_log.append("Creating fresh Chrome driver...")
        driver = _make_driver(headless=True)
        wait = WebDriverWait(driver, timeout)
        
        debug_log.append(f"Navigating to {RATEMYSITE_URL}")
        driver.get(RATEMYSITE_URL)
        
        debug_log.append("Checking for cookie banners...")
        _maybe_close_cookie_banner(driver)

        input_xpaths = [
            "//input[@type='text']",
            "//input[@placeholder]",
            "//input",
            "//textarea",
        ]
        
        debug_log.append("Looking for input field...")
        try:
            input_el = wait.until(EC.presence_of_element_located((By.XPATH, "|".join(input_xpaths))))
            debug_log.append("Found input field using wait condition")
        except Exception as e:
            debug_log.append(f"Wait condition failed: {e}")
            input_el = _find_first(driver, input_xpaths)
            if input_el:
                debug_log.append("Found input field using fallback method")
            
        if not input_el:
            debug_log.append("ERROR: Could not locate input field!")
            try:
                body_text = driver.find_element(By.TAG_NAME, "body").text[:500]
                debug_log.append(f"Body text: {body_text}")
            except Exception as e:
                debug_log.append(f"Could not get body text: {e}")
            return "", debug_log

        debug_log.append(f"Entering URL: {target_url}")
        try:
            input_el.clear()
        except Exception:
            pass
        input_el.send_keys(target_url)
        time.sleep(0.3)

        debug_log.append("Attempting to submit...")
        clicked = _click_best_button(driver)
        if clicked:
            debug_log.append("Successfully clicked submit button")
        else:
            debug_log.append("Button click failed, trying Enter key...")
            try:
                input_el.send_keys("\n")
                debug_log.append("Sent Enter key")
            except Exception as e:
                debug_log.append(f"Enter key failed: {e}")

        debug_log.append("Waiting for results to load...")
        try:
            wait.until(
                EC.presence_of_element_located((By.XPATH, "//span[contains(text(), 'Overall Score')]"))
            )
            debug_log.append("Found Overall Score element")
        except TimeoutException:
            debug_log.append("Overall Score not found, waiting for general content...")
            time.sleep(5)

        time.sleep(3)
        debug_log.append("Extracting result HTML...")
        result_html = _collect_result_html(driver)
        debug_log.append(f"Extracted {len(result_html)} characters of HTML")
        
        return result_html, debug_log

    except Exception as e:
        debug_log.append(f"ERROR in analysis: {e}")
        debug_log.append(f"Traceback: {traceback.format_exc()}")
        return "", debug_log
    finally:
        if driver:
            debug_log.append("Closing driver...")
            try:
                driver.quit()
            except Exception:
                pass

def _clean_text(text: str) -> str:
    """Clean text while preserving meaningful structure"""
    if not text or text == "-":
        return "-"
    
    text = re.sub(r'\s+', ' ', text.strip())
    text = re.sub(r'[^\w\s\.,;:()!?\-\'\"\/&]', '', text)
    
    return text.strip()

def _parse_ratemysite_html(html: str, url: str) -> Dict[str, str]:
    """Parse RateMySite HTML structure to extract scores and descriptions"""
    soup = BeautifulSoup(html, 'html.parser')
    
    company_name = url.replace("https://", "").replace("http://", "").replace("www.", "").split("/")[0]
    
    result = {
        "Company": company_name,
        "URL": url,
        "Overall Score": "-",
        "Description of Website": "-",
        "Consumer Score": "-",
        "Consumer Score Description": "-",
        "Developer Score": "-",
        "Developer Score Description": "-",
        "Investor Score": "-",
        "Investor Score Description": "-",
        "Clarity Score": "-",
        "Clarity Score Description": "-",
        "Visual Design Score": "-",
        "Visual Design Score Description": "-",
        "UX Score": "-",
        "UX Score Description": "-",
        "Trust Score": "-",
        "Trust Score Description": "-",
        "Value Prop Score": "-",
        "Value Prop Score Description": "-",
    }
    
    try:
        # Extract Overall Score
        overall_score_elem = soup.find('span', class_='text-5xl')
        if overall_score_elem:
            score_text = overall_score_elem.get_text(strip=True)
            if re.match(r'^\d+\.?\d*$', score_text):
                result["Overall Score"] = score_text
        
        # Extract main description
        desc_elem = soup.find('p', class_='text-xl text-white')
        if desc_elem:
            result["Description of Website"] = _clean_text(desc_elem.get_text(strip=True))
        
        # Extract Audience Perspective scores and descriptions
        audience_section = soup.find('h2', string='Audience Perspective')
        if audience_section:
            audience_container = audience_section.find_next('div', class_='grid')
            if audience_container:
                audience_cards = audience_container.find_all('div', recursive=False)
                
                for card in audience_cards:
                    title_elem = card.find('h3')
                    if not title_elem:
                        continue
                    
                    title = title_elem.get_text(strip=True)
                    score_elem = card.find('span', class_='text-2xl')
                    score = score_elem.get_text(strip=True) if score_elem else "-"
                    desc_elem = card.find('p', class_='text-gray-300')
                    description = _clean_text(desc_elem.get_text(strip=True)) if desc_elem else "-"
                    
                    if title == "Consumer":
                        result["Consumer Score"] = score
                        result["Consumer Score Description"] = description
                    elif title == "Developer":
                        result["Developer Score"] = score
                        result["Developer Score Description"] = description
                    elif title == "Investor":
                        result["Investor Score"] = score
                        result["Investor Score Description"] = description
        
        # Extract Technical Criteria scores and descriptions
        tech_section = soup.find('h2', string='Technical Criteria Scores')
        if tech_section:
            tech_container = tech_section.find_next('div', class_='grid')
            if tech_container:
                tech_cards = tech_container.find_all('div', class_='p-6')
                
                for card in tech_cards:
                    title_elem = card.find('h3')
                    if not title_elem:
                        continue
                    
                    title = title_elem.get_text(strip=True)
                    score_elem = card.find('span', class_='text-2xl')
                    score = score_elem.get_text(strip=True) if score_elem else "-"
                    desc_elem = card.find('p', class_='text-gray-300')
                    description = _clean_text(desc_elem.get_text(strip=True)) if desc_elem else "-"
                    
                    if title == "Clarity":
                        result["Clarity Score"] = score
                        result["Clarity Score Description"] = description
                    elif title == "Visual Design":
                        result["Visual Design Score"] = score
                        result["Visual Design Score Description"] = description
                    elif title == "UX":
                        result["UX Score"] = score
                        result["UX Score Description"] = description
                    elif title == "Trust":
                        result["Trust Score"] = score
                        result["Trust Score Description"] = description
                    elif title == "Value Proposition":
                        result["Value Prop Score"] = score
                        result["Value Prop Score Description"] = description
        
    except Exception as e:
        print(f"Error parsing HTML: {e}")
    
    return result

def create_excel_report(results_data: List[Dict[str, str]]) -> io.BytesIO:
    """Create a formatted Excel report from analysis results"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Competitor UI/UX Analysis"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    section_font = Font(bold=True, color="FFFFFF")
    section_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    score_font = Font(bold=True, color="0066CC")
    score_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    desc_font = Font(italic=True, color="5D6D7E")
    desc_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Add title and metadata
    ws.merge_cells('A1:' + get_column_letter(len(results_data) + 1) + '1')
    title_cell = ws['A1']
    title_cell.value = "Step 3: Competitor Web UI/UX Analysis Report"
    title_cell.font = Font(bold=True, size=16, color="2C3E50")
    title_cell.alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A2:' + get_column_letter(len(results_data) + 1) + '2')
    date_cell = ws['A2']
    date_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    date_cell.font = Font(size=10, color="5D6D7E")
    date_cell.alignment = Alignment(horizontal="center")
    
    current_row = 4
    
    # Headers
    ws.cell(row=current_row, column=1, value="UI/UX Metrics")
    ws.cell(row=current_row, column=1).font = header_font
    ws.cell(row=current_row, column=1).fill = header_fill
    ws.cell(row=current_row, column=1).border = border
    
    for col, result in enumerate(results_data, start=2):
        cell = ws.cell(row=current_row, column=col, value=result.get('Company', 'Unknown'))
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    current_row += 1
    
    # Define the metrics structure
    metrics = [
        ("Company", "Company", "basic"),
        ("URL", "Website URL", "basic"),
        ("Overall Score", "Overall UI/UX Score", "score"),
        ("Description of Website", "Website Overview", "desc"),
        ("", "👥 User Experience Analysis", "section"),
        ("Consumer Score", "Consumer Appeal Score", "score"),
        ("Consumer Score Description", "Consumer Experience Analysis", "desc"),
        ("Developer Score", "Technical Implementation Score", "score"),
        ("Developer Score Description", "Technical Assessment", "desc"),
        ("Investor Score", "Business Impact Score", "score"),
        ("Investor Score Description", "Business Value Analysis", "desc"),
        ("", "🎨 Design & Usability Metrics", "section"),
        ("Clarity Score", "Content Clarity Score", "score"),
        ("Clarity Score Description", "Content & Information Architecture", "desc"),
        ("Visual Design Score", "Visual Design Score", "score"),
        ("Visual Design Score Description", "Visual Design Assessment", "desc"),
        ("UX Score", "User Experience Score", "score"),
        ("UX Score Description", "UX & Navigation Analysis", "desc"),
        ("Trust Score", "Trust & Credibility Score", "score"),
        ("Trust Score Description", "Trust Indicators Assessment", "desc"),
        ("Value Prop Score", "Value Communication Score", "score"),
        ("Value Prop Score Description", "Value Proposition Clarity", "desc"),
    ]
    
    # Add metrics
    for key, label, metric_type in metrics:
        if metric_type == "section":
            ws.merge_cells(f'A{current_row}:' + get_column_letter(len(results_data) + 1) + f'{current_row}')
            section_cell = ws.cell(row=current_row, column=1, value=label)
            section_cell.font = section_font
            section_cell.fill = section_fill
            section_cell.alignment = Alignment(horizontal="center")
            section_cell.border = border
        else:
            label_cell = ws.cell(row=current_row, column=1, value=label)
            label_cell.border = border
            
            if metric_type == "score":
                label_cell.fill = score_fill
                label_cell.font = score_font
            elif metric_type == "desc":
                label_cell.fill = desc_fill
                label_cell.font = desc_font
            
            for col, result in enumerate(results_data, start=2):
                value = result.get(key, "-")
                data_cell = ws.cell(row=current_row, column=col, value=value)
                data_cell.border = border
                
                if metric_type == "score":
                    data_cell.fill = score_fill
                    if value != "-" and value.replace(".", "").isdigit():
                        data_cell.font = Font(bold=True, color="0066CC")
                        data_cell.alignment = Alignment(horizontal="center")
                elif metric_type == "desc":
                    data_cell.fill = desc_fill
                    data_cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        current_row += 1
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 30
    for col in range(2, len(results_data) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 25
    
    # Set row heights for description rows
    for row in range(1, current_row):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value and ("Description" in str(cell_value) or "Analysis" in str(cell_value) or "Assessment" in str(cell_value)):
            ws.row_dimensions[row].height = 50
    
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file

app = Flask(__name__)

def sse(event: str, data: dict) -> str:
    return f"event: {event}\ndata: {json.dumps(data)}\n\n"

def stream_analysis(urls: List[str]) -> Generator[str, None, None]:
    global analysis_results
    analysis_results = {}
    
    total = len(urls)
    yield sse("init", {"total": total})

    for idx, raw in enumerate(urls, start=1):
        url = raw if raw.startswith(("http://", "https://")) else "https://" + raw
        step_total = 5

        print(f"[{idx}/{total}] Start {url}")
        yield sse("start_url", {"index": idx, "url": url})

        yield sse("progress", {"index": idx, "phase": "Creating fresh browser", "p": 1, "of": step_total})
        yield sse("progress", {"index": idx, "phase": "Submitting to RateMySite", "p": 2, "of": step_total})
        
        raw_html, debug_messages = _analyze_one_with_debugging(url, timeout=DEFAULT_TIMEOUT)
        
        for msg in debug_messages:
            yield sse("debug", {"index": idx, "message": msg})

        yield sse("progress", {"index": idx, "phase": "Parsing output", "p": 3, "of": step_total})
        
        if raw_html:
            data = _parse_ratemysite_html(raw_html, url)
            analysis_results[url] = data
            yield sse("result", {"index": idx, "url": url, "data": data})
        else:
            error_data = {"Company": "Analysis Failed", "URL": url, "Overall Score": "-"}
            analysis_results[url] = error_data
            yield sse("result", {"index": idx, "url": url, "error": "No results found - check debug log"})

        yield sse("progress", {"index": idx, "phase": "Done", "p": step_total, "of": step_total})
        print(f"[{idx}/{total}] Done {url}")

    yield sse("done", {"ok": True})

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/stream")
def stream():
    urls = [u.strip() for u in request.args.getlist("u") if u.strip()]
    if not urls:
        return Response("Need at least one ?u=", status=400)
    return Response(stream_with_context(stream_analysis(urls)), mimetype="text/event-stream")

@app.route("/download-excel")
def download_excel():
    global analysis_results
    
    if not analysis_results:
        return jsonify({"error": "No analysis results available"}), 400
    
    try:
        results_list = list(analysis_results.values())
        excel_file = create_excel_report(results_list)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Competitor_UIUX_Analysis_{timestamp}.xlsx"
        
        return send_file(
            excel_file,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        print(f"Error creating Excel file: {e}")
        return jsonify({"error": "Failed to create Excel file"}), 500

@app.route("/health")
def health():
    return {"status": "healthy", "message": "Step 3: Competitor Web UI UX Analysis is running"}

if __name__ == "__main__":
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_ENV') == 'development'
    app.run(debug=debug, host="0.0.0.0", port=port)
